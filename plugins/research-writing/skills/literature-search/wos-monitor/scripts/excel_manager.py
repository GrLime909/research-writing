"""Excel read/append/dedup manager for journal-monitor skill."""

import re
from pathlib import Path
from typing import Optional
from dataclasses import dataclass, field
from datetime import datetime

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl required. Install: pip install openpyxl")

COLUMNS = [
    "论文名", "作者", "发表时间", "期刊",
    "DOI", "URL", "摘要",
    "论文中文名", "摘要中文翻译", "入库时间"
]


@dataclass
class PaperRecord:
    title: str
    authors: str
    pub_date: str
    journal: str
    doi: str
    url: str
    abstract: str
    title_cn: str = ""
    abstract_cn: str = ""
    added_time: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M"))

    def dedup_key(self) -> str:
        if self.doi:
            return f"doi:{self.doi.lower().strip()}"
        normalized = re.sub(r"\s+", " ", self.title.lower().strip())
        normalized = re.sub(r"[^\w\s]", "", normalized)
        return f"combo:{normalized}|{self.journal.lower().strip()}|{self.pub_date[:4]}"


class ExcelManager:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.existing_keys: set[str] = set()

    def load_existing_keys(self) -> set[str]:
        if not self.path.exists():
            return set()
        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        doi_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value and "DOI" in str(cell.value):
                doi_col = col_idx
                break
        if doi_col is None:
            wb.close()
            return set()
        keys = set()
        title_col = doi_col - 3    # 论文名
        journal_col = doi_col - 1  # 期刊
        date_col = doi_col - 2     # 发表时间
        for row in ws.iter_rows(min_row=2, values_only=True):
            doi_val = row[doi_col - 1] if len(row) >= doi_col else None
            if doi_val and str(doi_val).strip():
                keys.add(f"doi:{str(doi_val).strip().lower()}")
            else:
                t = row[title_col - 1] if len(row) >= title_col else None
                j = row[journal_col - 1] if len(row) >= journal_col else None
                d = row[date_col - 1] if len(row) >= date_col else None
                if t and j:
                    normalized = re.sub(r"\s+", " ", str(t).lower().strip())
                    normalized = re.sub(r"[^\w\s]", "", normalized)
                    yr = str(d)[:4] if d else "0000"
                    keys.add(f"combo:{normalized}|{str(j).lower().strip()}|{yr}")
        wb.close()
        self.existing_keys = keys
        return keys

    def is_new(self, paper: PaperRecord) -> bool:
        return paper.dedup_key() not in self.existing_keys

    def deduplicate_batch(self, papers: list[PaperRecord]) -> tuple[list[PaperRecord], list[PaperRecord]]:
        seen = set(self.existing_keys)
        new_papers = []
        duplicates = []
        for p in papers:
            key = p.dedup_key()
            if key in seen:
                duplicates.append(p)
            else:
                seen.add(key)
                new_papers.append(p)
        return new_papers, duplicates

    def append(self, papers: list[PaperRecord]):
        if not papers:
            return
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(COLUMNS)
        for p in papers:
            ws.append([
                p.title, p.authors, p.pub_date, p.journal,
                p.doi, p.url, p.abstract,
                p.title_cn, p.abstract_cn, p.added_time
            ])
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                ws.column_dimensions[get_column_letter(col_idx)].width or 8,
                min(60, max(8, max((len(str(cell.value or "")) for cell in ws[get_column_letter(col_idx)])) + 2))
            )
        wb.save(self.path)
        wb.close()
        for p in papers:
            self.existing_keys.add(p.dedup_key())
