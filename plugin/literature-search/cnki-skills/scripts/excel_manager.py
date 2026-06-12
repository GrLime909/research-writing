"""Excel read/append/dedup manager for cnki-skills — shared with WOS.

Columns: 文章标题, 作者, 所属期刊, 摘要, 发表时间, 中文标题, 中文摘要, 入库时间
Dedup: DOI first, then (title + journal + year) fallback.
"""

import json
import re
import sys
from pathlib import Path
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

COLUMNS = [
    "文章标题", "作者", "所属期刊", "摘要", "发表时间",
    "中文标题", "中文摘要", "入库时间"
]


@dataclass
class PaperRecord:
    title: str
    authors: str
    journal: str
    abstract: str
    pub_date: str
    doi: str = ""
    title_cn: str = ""
    abstract_cn: str = ""
    added_time: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M"))

    def dedup_key(self) -> str:
        if self.doi and self.doi.strip():
            return f"doi:{self.doi.strip().lower()}"
        normalized = re.sub(r"\s+", " ", self.title.lower().strip())
        normalized = re.sub(r"[^\w\s]", "", normalized)
        yr = re.search(r"(\d{4})", self.pub_date)
        yr_str = yr.group(1) if yr else "0000"
        return f"combo:{normalized}|{self.journal.lower().strip()}|{yr_str}"

    def to_row(self) -> list:
        return [
            self.title, self.authors, self.journal, self.abstract,
            self.pub_date, self.title_cn, self.abstract_cn, self.added_time
        ]


class ExcelManager:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.existing_keys: set[str] = set()

    def load_existing_keys(self) -> set[str]:
        if not self.path.exists():
            return set()
        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        # Find column indices from header
        col_map = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                col_map[str(cell.value).strip()] = col_idx
        wb.close()

        doi_col = col_map.get("DOI") or col_map.get("DOI")
        # Re-open for reading values (read_only mode can't iterate twice easily)
        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            row_vals = [str(c) if c else "" for c in row]
            # Check DOI column if we found it
            doi_val = ""
            if doi_col and doi_col <= len(row_vals):
                doi_val = row_vals[doi_col - 1].strip()
            if doi_val and doi_val.lower().startswith("10."):
                keys.add(f"doi:{doi_val.lower()}")
                continue
            # Fallback: title (col 1) + journal (col 3) + pub_date (col 5)
            t = row_vals[0] if len(row_vals) > 0 else ""
            j = row_vals[2] if len(row_vals) > 2 else ""
            d = row_vals[4] if len(row_vals) > 4 else ""
            if t and j:
                normalized = re.sub(r"\s+", " ", t.lower().strip())
                normalized = re.sub(r"[^\w\s]", "", normalized)
                yr = re.search(r"(\d{4})", d)
                yr_str = yr.group(1) if yr else "0000"
                keys.add(f"combo:{normalized}|{j.lower().strip()}|{yr_str}")
        wb.close()
        self.existing_keys = keys
        return keys

    def deduplicate_batch(self, papers: list[PaperRecord]) -> tuple[list[PaperRecord], list[PaperRecord]]:
        seen = set(self.existing_keys)
        new_papers = []
        duplicates = []
        for p in papers:
            key = p.dedup_key()
            if key in seen:
                duplicates.append(p)
            else:
                seen.add(key)
                new_papers.append(p)
        return new_papers, duplicates

    def append(self, papers: list[PaperRecord]):
        if not papers:
            return
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(COLUMNS)
        for p in papers:
            ws.append(p.to_row())
        # Auto-size columns
        for col_idx in range(1, len(COLUMNS) + 1):
            max_len = 8
            for cell in ws[get_column_letter(col_idx)]:
                if cell.value:
                    max_len = max(max_len, min(60, len(str(cell.value)) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len
        wb.save(self.path)
        wb.close()
        for p in papers:
            self.existing_keys.add(p.dedup_key())


def load_papers_from_json(data) -> list[PaperRecord]:
    """Load papers from JSON (list of dicts or single dict). Handles CNKI and WOS formats."""
    if isinstance(data, dict):
        if "items" in data:
            data = data["items"]
        elif "results" in data:
            data = data["results"]
        else:
            data = [data]
    papers = []
    for item in data:
        papers.append(PaperRecord(
            title=item.get("title", ""),
            authors=item.get("authors", "") if isinstance(item.get("authors"), str) else "; ".join(item.get("authors", [])),
            journal=item.get("journal", "") or item.get("source", "") or item.get("publicationTitle", ""),
            abstract=item.get("abstract", "") or item.get("abstractNote", ""),
            pub_date=item.get("date", "") or item.get("pub_date", "") or item.get("pubTime", ""),
            doi=item.get("doi", "") or item.get("DOI", ""),
            title_cn=item.get("title_cn", ""),
            abstract_cn=item.get("abstract_cn", ""),
        ))
    return papers


def main():
    import argparse
    p = argparse.ArgumentParser(description="CNKI/WOS Excel Manager")
    p.add_argument("--output", required=True, help="Path to output Excel (.xlsx)")
    p.add_argument("--input", default=None, help="JSON file with paper data (stdin if not provided)")
    p.add_argument("--action", default="append", choices=["append", "list", "dedup-check"])
    args = p.parse_args()

    excel = ExcelManager(args.output)

    if args.action == "list":
        excel.load_existing_keys()
        print(f"Existing papers: {len(excel.existing_keys)}")
        return

    # Read input
    if args.input:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = json.load(sys.stdin)

    papers = load_papers_from_json(data)

    if args.action == "dedup-check":
        excel.load_existing_keys()
        for p in papers:
            key = p.dedup_key()
            status = "DUPLICATE" if key in excel.existing_keys else "NEW"
            print(f"[{status}] {p.title[:60]} | {p.journal} | {p.pub_date}")
        return

    # append
    excel.load_existing_keys()
    new_papers, duplicates = excel.deduplicate_batch(papers)
    print(f"Already in Excel: {len(duplicates)}")
    print(f"New papers: {len(new_papers)}")
    if new_papers:
        excel.append(new_papers)
        print(f"Saved {len(new_papers)} papers to {args.output}")
        for p in new_papers:
            print(f"  + {p.title[:60]} | {p.journal} | {p.pub_date}")


if __name__ == "__main__":
    main()
